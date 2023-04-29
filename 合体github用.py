import openpyxl
import string
from PIL import Image
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from tqdm import tqdm



num = 0

# エクセルのpathを指定
excel_path = r"C:\Excekのパスを入力.xlsx"
# 読み込む画像のpathを指定
input_path = r"\読み込む画像のパスを入力.png"
# 書き込む画像のpathを指定
output_path = r"C:\読み込んだ画像をリサイズしたのちのパス.png"


def get_column_name(column_number):
    letters = []
    while column_number > 0:
        column_number, remainder = divmod(column_number - 1, 26)
        letters.append(chr(65 + remainder))
    return ''.join(reversed(letters))


def cell_name(x, y):
    col_num = x + 1
    quotient = col_num // 26
    remainder = col_num % 26
    if remainder == 0:
        remainder = 26
        quotient -= 1
    col_name = get_column_name(remainder)
    if quotient > 0:
        col_name = get_column_name(quotient) + col_name
    row_num = y + 1
    return f"{col_name}{row_num}"


def excel_print(x, y, rgba):
    # Excelファイルの読み込み
    workbook = openpyxl.load_workbook(excel_path)

    # セルの背景色を設定するrgba値
    rgba_values = (rgba)

    # rgba値を16進数に変換
    hex_value = '{:02X}{:02X}{:02X}'.format(*rgba_values[:3])

    # セルの背景色を設定する
    fill = PatternFill(start_color='FF' + hex_value, end_color='FF' + hex_value, fill_type='solid')
    sheet = workbook.active
    cell = sheet[cell_name(x,y)]
    cell.fill = fill

    # Excelファイルを保存
    workbook.save(excel_path)


# 画像ファイルを開く
image = Image.open(input_path)

# 画像をリサイズする
resized_image = image.resize((360, 360))  # 解像度はお好きなようにしてください。高画質にするほど時間がかかります。

# リサイズした画像を保存する
resized_image.save(output_path)

# 画像ファイルを開く
image = Image.open(output_path)

# 画像のピクセル情報を取得する
pixels = image.load()

# 画像の幅と高さを取得する
width, height = image.size

# 全てのピクセルのRGBA値を取得する
rgba_values = []
for y in tqdm(range(height)):
    for x in range(width):
        rgba = pixels[x, y]
        rgba_values.append(rgba)

print("画像の解析終了")

for y in tqdm(range(height)):
    for x in tqdm(range(width)):
        excel_print(x, y, rgba_values[num])
        num += 1